import PIL
import xlrd
import openpyxl
import tkinter
from tkinter.ttk import Combobox
from tkinter import *
from  tkinter import messagebox
from datetime import date, datetime
from tkinter import filedialog
from PIL import Image,ImageTk
import os
from tkinter.ttk import Combobox
from openpyxl import Workbook
import pathlib
root=Tk()
root.title('Management System')
root.geometry("1400x1250")
icon=Image.open("D://karim (2).png")
icon_r=icon.resize((200,200))
icons=ImageTk.PhotoImage(icon_r)

root.iconphoto(False,icons)
img=Image.open('D://kk1.png')
img=img.resize((1400,1250))
photo=ImageTk.PhotoImage(img)
lbl=Label(root,image=photo,width=1400,height=1250)
lbl.place(x=0,y=0)



#root.configure(bg='#dfd4d4')#917171  '#dfd4d4'#f0e6db
Label(root,bg='#efbf8e',fg='black',text='العملاء بيانات',font=('calibri',30,'italic')).place(x=600,y=20)
frame=Frame(root,width=1000,height=600,bg='#f0e6db')
frame.place(x=125,y=80)
file=pathlib.Path("D://customer1.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="رقم العميل"
    sheet['B1']="اسم العميل"
    sheet['C1']="رقم الهاتف"
    sheet['D1']="الدفعة الاولي"
    sheet['E1']="الدفعة التانيه"
    sheet['F1']="الدفعة التالته"
    sheet['G1']="الدفعة الرابعة"
    sheet['H1']="الدفعة الخامسة"
    sheet['I1']="اجمالي المبلغ"
    sheet['J1']="المبلغ المتبق"
    sheet['K1']="صاحب الرخام"
    sheet['L1']="رقم الرخام"
    sheet['M1']='تكلفة الرخام'
    sheet['N1']='المدفوع'
    sheet['O1']="باقي الرخام"

    file.save('D://customer1.xlsx')
#################registration number#####################
def registraion():
    file = openpyxl.load_workbook("D://customer1.xlsx")
    sheet = file.active

    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value
    try:
        reg_no.set(max_row_value + 1)
    except:
        reg_no.set("1")
####################search########################
def find():
    text = search.get()
    reset()
    file=openpyxl.load_workbook('D://customer1.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[2].value==str(text):
            name=row[2]
            reg_no_pos=str(name)[14:-1]
            reg_no9=str(name)[15:-1]
            try:
                print(str(name))
            except:
                messagebox.showerror('error',"invalid registration no!!!!!!")
            x1=sheet.cell(row=int(reg_no9),column=1).value
            x2=sheet.cell(row=int(reg_no9),column=2).value
            x3 = sheet.cell(row=int(reg_no9), column=3).value
            x4 = sheet.cell(row=int(reg_no9), column=4).value
            x5 = sheet.cell(row=int(reg_no9), column=5).value
            x6 = sheet.cell(row=int(reg_no9), column=6).value
            x7 = sheet.cell(row=int(reg_no9), column=7).value
            x8 = sheet.cell(row=int(reg_no9), column=8).value
            x9 = sheet.cell(row=int(reg_no9), column=9).value
            x10 = sheet.cell(row=int(reg_no9), column=10).value
            x11 = sheet.cell(row=int(reg_no9), column=11).value
            x12=sheet.cell(row=int(reg_no9), column=12).value
            x13=sheet.cell(row=int(reg_no9),column=13).value
            x14 = sheet.cell(row=int(reg_no9), column=14).value
            x15=sheet.cell(row=int(reg_no9),column=15).value
            reg_no.set(x1)
            username.set(x2)
            phone.set(x3)
            first.set(x4)
            second.set(x5)
            third.set(x6)
            fourth.set(x7)
            five.set(x8)
            total_cost.set(x9)
            paid.set(x10)
            r_name.set(x11)
            r_phone.set(x12)
            r_total.set(x13)
            r_paid.set(x14)
            total_price.set(x15)
def save():
    r = reg_no.get()
    f = username.get()
    p = phone.get()
    to = total_cost.get()
    f1 = first.get()
    s = second.get()
    th = third.get()
    f4 = fourth.get()
    f5 = five.get()
    p2 = paid.get()
    r1 = r_name.get()
    r2 = r_phone.get()
    r_t=r_total.get()
    r_p=r_paid.get()
    r_pt=total_price.get()
    if r == "" or f== "" :

     messagebox.showerror("error", "few data is missing")
    else:
            file = openpyxl.load_workbook("D://customer1.xlsx")
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row + 1, value=r)
            sheet.cell(column=2, row=sheet.max_row, value=f)
            sheet.cell(column=3, row=sheet.max_row, value=p)
            sheet.cell(column=4, row=sheet.max_row, value=f1)
            sheet.cell(column=5, row=sheet.max_row, value=s)
            sheet.cell(column=6, row=sheet.max_row, value=th)
            sheet.cell(column=7, row=sheet.max_row, value=f4)
            sheet.cell(column=8, row=sheet.max_row, value=f5)
            sheet.cell(column=9, row=sheet.max_row, value=to)
            sheet.cell(column=10, row=sheet.max_row, value=p2)
            sheet.cell(column=11, row=sheet.max_row, value=r1)
            sheet.cell(column=12, row=sheet.max_row, value=r2)
            sheet.cell(column=13,row=sheet.max_row,value=r_t)
            sheet.cell(column=14,row=sheet.max_row,value=r_p)
            sheet.cell(column=15,row=sheet.max_row,value=r_pt)
            file.save(r'D://customer1.xlsx')
            messagebox.showinfo('info','تم اضافة العميل')
            reset()
#######################paid###################
def pay():
        global lbl_entry

        try:
            a1 = int(first.get())
        except:
            a1 = 0
        try:
            a2 = int(second.get())
        except:
            a2 = 0
        try:
            a3 = int(third.get())
        except:
            a3 = 0
        try:
            a4 = int(fourth.get())
        except:
            a4 = 0
        try:
            a5 = int(five.get())
        except:
            a5 = 0
        try:
            a6 = int(total_cost.get())
        except:
            a6 = 0
        t=a1+a2+a3+a4+a5



        lbl_entry = Entry(frame, font=('arial', 20, 'bold'), textvariable=paid, bd=2, width=5, bg='lightblue')
        lbl_entry.place(x=380, y=253)
        total_cost1 = a6-t

        paid.set(total_cost1)
#######################r_paid################
def total1():


    try:
        y1 = int(r_total.get())
    except:
        y1 = 0
    try:
        y2 = int(r_paid.get())
    except:
        y2 = 0

    x = y1 - y2

    lbl1_entry = Entry(f2, font=('arial', 20, 'bold'), textvariable=total_price, bd=2, width=5, bg='lightblue')
    lbl1_entry.place(x=200, y=173)


    total_price.set(x)
#################reset############

def reset():
    username_entry.delete(0,END)
    phone_entry.delete(0,END)
    first_entry.delete(0,END)
    second_entry.delete(0,END)
    third_entry.delete(0,END)
    fourth_entry.delete(0,END)
    five_entry.delete(0,END)
    total_cost_entry.delete(0,END)
    r_name_entry.delete(0,END)
    r_phone_entry.delete(0,END)

    r_total_entry.delete(0,END)
    r_paid_entry.delete(0,END)
    total_price.set("")
def exit():
    root.destroy()
############update############
def update():
    r = reg_no.get()
    f = username.get()
    p = phone.get()
    to = total_cost.get()
    f1 = first.get()
    s = second.get()
    th = third.get()
    f4 = fourth.get()
    f5 = five.get()
    p2 = paid.get()
    r1 = r_name.get()
    r2 = r_phone.get()
    r_t=r_total.get()
    r_p=r_paid.get()
    r_pt=total_price.get()
    file=openpyxl.load_workbook('D://customer1.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(r):
            Name = row[0]
            reg_no_pos = str(Name)[14:-1]
            reg_no9 = str(Name)[15:-1]
            try:
                print(str(Name))
            except:
                messagebox.showerror('error', "invalid registration no!!!!!!")
            sheet.cell(row=int(reg_no9), column=1,value=r)
            sheet.cell(row=int(reg_no9), column=2,value=f)
            sheet.cell(row=int(reg_no9), column=3,value=p)
            sheet.cell(row=int(reg_no9), column=4,value=to)
            sheet.cell(row=int(reg_no9), column=5,value=f1)
            sheet.cell(row=int(reg_no9), column=6,value=s)
            sheet.cell(row=int(reg_no9), column=7,value=th)
            sheet.cell(row=int(reg_no9), column=8,value=f4)
            sheet.cell(row=int(reg_no9), column=9,value=f5)
            sheet.cell(row=int(reg_no9), column=10,value=p2)
            sheet.cell(row=int(reg_no9), column=11,value=r1)
            sheet.cell(row=int(reg_no9), column=12,value=r2)
            sheet.cell(row=int(reg_no9), column=13, value=r_t)
            sheet.cell(row=int(reg_no9), column=14, value=r_p)
            sheet.cell(row=int(reg_no9), column=15, value=r_pt)


            file.save('D://customer1.xlsx')
            messagebox.showinfo("info", "تم تعديل البيانات بنجاح!!!!")
            reset()

f1=Label(frame,width=50,height=600,bg='#8e9e88')
f1.place(x=0,y=0)
#################################

Label(frame,text="رقم العميل",font=('tahoma',14,'bold'),bg='#f0e6db').place(x=890,y=50)
reg_no=IntVar()
reg_entry=Entry(frame,textvariable=reg_no,font=('tahoma',14,'bold'),width=15,bg='white',fg='black')
reg_entry.place(x=700,y=50)
registraion()
##########################################
Label(frame, text="اسم العميل", font=('tahoma',14,'bold'), bg='#f0e6db',justify='right').place(x=560, y=50)
username=StringVar()
username_entry=Entry(frame,width=15,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=username,justify="right")

username_entry.place(x=370,y=50)
#############################################
Label(frame,text="رقم الجوال",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=890,y=120)
phone=StringVar()
phone_entry=Entry(frame,width=15,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=phone,justify="right")
phone_entry.place(x=700,y=120)
############################################
Label(frame,text="اجمالي المبلغ",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=560,y=120)
total_cost=StringVar()
total_cost_entry=Entry(frame,width=15,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=total_cost,justify="right")
total_cost_entry.place(x=370,y=120)
###########################################
Label(frame,text="الدفعة الاولي",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=870,y=190)
first=StringVar()
first_entry=Entry(frame,width=5,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=first,justify="right")
first_entry.place(x=800,y=190)


##########################################
Label(frame,text="الدفعة التانيه",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=660,y=190)
second=StringVar()
second_entry=Entry(frame,width=5,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=second,justify="right")
second_entry.place(x=580,y=190)
########################################
Label(frame,text="الدفعة التالته",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=440,y=190)
third=StringVar()
third_entry=Entry(frame,width=5,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=third,justify="right")
third_entry.place(x=370,y=190)
########################################
Label(frame,text="الدفعة الرابعه",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=870,y=260)
fourth=StringVar()
fourth_entry=Entry(frame,width=5,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=fourth,justify="right")
fourth_entry.place(x=800,y=260)
##########################################
Label(frame,text="الدفعة الخامسه",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=640,y=260)
five=StringVar()
five_entry=Entry(frame,width=5,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=five,justify="right")
five_entry.place(x=570,y=260)
########################################


paid=StringVar()
btn_total=Button(frame,bg='lightblue',fg='black',font=('tahoma',14,'bold'),width=5,height=1,text="الباقي",command=pay,justify="right")
btn_total.place(x=470,y=250)
######################################
f2=LabelFrame(frame,width=700,bg='#f0e6db',height=350,text='الرخام',font=('tahoma',20,'bold'))
f2.place(x=358,y=350)
Label(f2,text="الاسم",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=580,y=30)
r_name=StringVar()
r_name_entry=Entry(f2,textvariable=r_name,font=('tahoma',14,'bold'),width=15,bg='white',fg='black',justify="right")
r_name_entry.place(x=365,y=30)

##########################################
Label(f2,text="رقم الجوال",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=240,y=30)
r_phone=StringVar()
r_phone_entry=Entry(f2,width=15,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=r_phone,justify="right")
r_phone_entry.place(x=30,y=30)
###########################################
Label(f2,text="التكلفه",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=570,y=100)
r_total=StringVar()
r_total_entry=Entry(f2,textvariable=r_total,font=('tahoma',14,'bold'),width=15,bg='white',fg='black',justify="right")
r_total_entry.place(x=365,y=100)

##########################################
Label(f2,text="المدفوع",font=('tahoma',14,'bold'),bg='#f0e6db',justify="right").place(x=240,y=100)
r_paid=StringVar()
r_paid_entry=Entry(f2,width=15,bg='white',fg='black',font=('tahoma',14,'bold'),textvariable=r_paid,justify="right")
r_paid_entry.place(x=30,y=100)
#######################################
total_price=StringVar()
btn_total_price=Button(f2,bg='lightblue',fg='black',font=('tahoma',14,'bold'),width=5,height=1,text="الباقي",command=total1,justify="right")
btn_total_price.place(x=300,y=170)
####################################

Label(f1,text="رقم الجوال",fg="black",bg="#8e9e88",font=('tahoma',12,'bold')).place(x=260,y=50)
search=StringVar()
Entry(f1,textvariable=search,font="arial 20 ",width=11).place(x=88,y=50)
image=Image.open('D://search.png')
image=image.resize((30,30))
photo2=ImageTk.PhotoImage(image)
src=Button(f1,text="بحث",compound=RIGHT,image=photo2,width=70,bg="#b7cfd4",font=('tahoma',14,'bold'),command=find)
src.place(x=7,y=50)
##############################################
save_btn=Button(f1,width=15,height=2,text=' اضافة  ',bg='#82c0cc',fg='white',font=('tahoma',14,'bold'),command=save)
save_btn.place(x=50,y=150)
update_btn=Button(f1,width=15,height=2,text="تعديل",bg='#efbf8e',fg='white',font=('tahoma',14,'bold'),command=update)
update_btn.place(x=50,y=250)
delete_btn=Button(f1,width=15,height=2,text="حذف",bg='#999999',fg='white',font=('tahoma',14,'bold'),command=reset)
delete_btn.place(x=50,y=350)
exit_btn=Button(f1,width=15,height=2,text="خروج",bg='#cc0000',fg='white',font=('tahoma',14,'bold'),command=exit)
exit_btn.place(x=50,y=450)

root.mainloop()
